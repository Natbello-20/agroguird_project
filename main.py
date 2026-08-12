"""
AgroGuard - Phase 2 Backend
With real OpenWeatherMap weather integration
"""

import json
import os
import random
import time
import uvicorn
import httpx
from pathlib import Path
from typing import Optional, Dict, Any
from datetime import datetime, timedelta

# Load environment
from dotenv import load_dotenv
load_dotenv()

# FastAPI imports
from fastapi import FastAPI, Request, UploadFile, File, Header, Query, Form, HTTPException, status, Response, Depends
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware

# Local imports
import database
import auth
import schemas


# ============================================================================
# SETUP
# ============================================================================

app = FastAPI(title="AgroGuard", version="1.0.0")

# Session Middleware (must be added before other middleware)
app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("JWT_SECRET_KEY", "your-secret-key-change-in-production-12345678"),
    session_cookie="agroguard_session",
    max_age=86400 * 7,  # 7 days
    same_site="lax",
    https_only=False  # Set to True in production with HTTPS
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Templates
templates = Jinja2Templates(directory="templates")

# Static Files
from fastapi.staticfiles import StaticFiles
app.mount("/static", StaticFiles(directory="static"), name="static")

# Config
HOST = os.getenv("HOST", "0.0.0.0")
PORT = int(os.getenv("PORT", "8000"))
DEBUG = os.getenv("DEBUG", "False").lower() == "true"

# ============================================================================
# DISEASE MODEL
# ============================================================================

# Import the disease detection model
from model import DiseaseDetectionModel

# Initialize model - FORCE REAL MODEL ONLY (No mock mode confusion)
# The real model is at: mobile_assets/maize_model.tflite
print("[INIT] Initializing disease detection model...")
try:
    model = DiseaseDetectionModel(model_path="mobile_assets/maize_model.tflite")
    print(f"✅ [MODEL] Real TensorFlow model loaded successfully!")
    print(f"✅ [MODEL] Input size: {model.image_size}, Classes: {len(model.labels)}")
except Exception as e:
    print(f"❌ [MODEL] CRITICAL ERROR: Failed to load TensorFlow model!")
    print(f"❌ [MODEL] Error: {e}")
    print(f"❌ [MODEL] Application cannot start without the model.")
    raise  # Stop the application - don't run without model

# ============================================================================
# ROUTES
# ============================================================================

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    """Landing page - Role selection"""
    return templates.TemplateResponse("role_selection.html", {"request": request})

@app.get("/farmer", response_class=HTMLResponse)
async def farmer_app(request: Request):
    """Farmer mobile app"""
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/about", response_class=HTMLResponse)
async def about_page(request: Request):
    """About AgroGuard page"""
    return templates.TemplateResponse("about.html", {"request": request})

@app.get("/privacy", response_class=HTMLResponse)
async def privacy_page(request: Request):
    """Privacy Policy page"""
    return templates.TemplateResponse("privacy_policy.html", {"request": request})

@app.post("/predict")
async def predict_disease(
    file: UploadFile = File(...),
    lang: str = Query("en"),
    device_id: str = Header(None),  # Anonymous Farmer ID
    x_latitude: str = Header(None),  # GPS latitude from phone
    x_longitude: str = Header(None),  # GPS longitude from phone
):
    try:
        # Validate required headers for production use
        if not device_id:
            # Allow testing without device_id but log warning
            print("⚠ Warning: No device_id provided. Scan limiting disabled.")
            device_id = f"anonymous_{int(time.time())}"
        
        # Compute segment identifier from GPS (rounded to 4 decimal places = ~11 meters)
        segment_id: Optional[str] = None
        if x_latitude and x_longitude:
            try:
                lat = round(float(x_latitude), 4)
                lon = round(float(x_longitude), 4)
                segment_id = f"{lat}_{lon}"
                print(f"✓ GPS Segment: {segment_id}")
            except ValueError:
                print("⚠ Warning: Invalid GPS coordinates")
                segment_id = None
        else:
            print("⚠ Warning: No GPS coordinates provided. Location-based limiting disabled.")

        # Validate file type
        if not file.content_type.startswith("image/"):
            return JSONResponse({
                "error": "Invalid file type. Please upload an image.",
                "disease": "Invalid File",
                "confidence": 0,
                "treatment": "",
                "recommendations": []
            }, status_code=400)
        
        # Read image bytes
        contents = await file.read()
        
        # Check size (5MB)
        if len(contents) > 5 * 1024 * 1024:
            return JSONResponse({
                "error": "Image too large (max 5MB)",
                "disease": "File Too Large",
                "confidence": 0,
                "treatment": "",
                "recommendations": []
            }, status_code=400)

        # Convert bytes to OpenCV image
        import numpy as np
        import cv2
        nparr = np.frombuffer(contents, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return JSONResponse({
                "error": "Could not decode image. Please try another image.",
                "disease": "Invalid Image",
                "confidence": 0,
                "treatment": "",
                "recommendations": []
            })

        # GREEN PRE-FILTER REMOVED: Model now has "not_maize" class and can detect non-maize objects!
        # The retrained model is smart enough to reject fabric, hands, tables, etc.
        # We rely on the AI model's intelligence instead of simple color filtering.

        # Run prediction with the NEW retrained model (now has "not_maize" class!)
        print(f"[PREDICT] Starting prediction with real TensorFlow model...")
        print(f"[PREDICT] Model status - model_loaded: {model.model_loaded}")
        
        result = model.predict(img)
        disease, confidence = result[0], result[1]
        entropy = result[2] if len(result) > 2 else 0.0
        confidence_gap = result[3] if len(result) > 3 else 0.0
        
        print(f"[PREDICT] Result - Disease: {disease}, Confidence: {confidence:.4f}, Entropy: {entropy:.4f}, Gap: {confidence_gap:.4f}")
        
        if disease is None:
            return JSONResponse({
                "error": "Prediction failed. Please try again.",
                "disease": "Analysis Failed",
                "confidence": 0,
                "treatment": "",
                "recommendations": []
            })
        
        # NEW: Check if model predicted "Not_Maize" class
        if disease == "Corn___Not_Maize":
            print(f"[MODEL-REJECT] Model detected non-maize object with confidence: {confidence:.4f}")
            return JSONResponse({
                "error": "This is not a maize leaf. Please capture a maize leaf to check for diseases.",
                "disease": "Not Maize Leaf",
                "confidence": float(round(confidence, 2)),
                "treatment": "",
                "recommendations": [],
                "debug_info": {
                    "rejection_reason": "model_predicted_not_maize",
                    "confidence": float(round(confidence, 4)),
                    "entropy": float(round(entropy, 4)),
                    "confidence_gap": float(round(confidence_gap, 4))
                }
            }, status_code=400)
        
        # Additional validation for maize classes (confidence threshold)
        # Even if model says it's maize, check if it's confident enough
        
        CONFIDENCE_THRESHOLD = 0.55  # Lowered from 0.60 since model now has not_maize class
        
        if confidence < CONFIDENCE_THRESHOLD:
            print(f"[REJECT] Low confidence - model uncertain: {confidence:.4f} < {CONFIDENCE_THRESHOLD}")
            return JSONResponse({
                "error": "Image quality is too poor or unclear. Please try again with better lighting.",
                "disease": "Poor Image Quality",
                "confidence": float(round(confidence, 2)),
                "treatment": "",
                "recommendations": [],
                "debug_info": {
                    "rejection_reason": f"low confidence ({confidence:.2f} < {CONFIDENCE_THRESHOLD})",
                    "confidence": float(round(confidence, 4)),
                    "entropy": float(round(entropy, 4)),
                    "confidence_gap": float(round(confidence_gap, 4))
                }
            }, status_code=400)
        
        # Additional safety: Verify it's a corn disease (model should only output Corn___ classes)
        if not disease.startswith('Corn___'):
            return JSONResponse({
                "error": "This is not a maize leaf. Please capture a maize leaf to check for diseases.",
                "disease": "Not Maize Leaf",
                "confidence": float(round(confidence, 2)),
                "treatment": "",
                "recommendations": []
            }, status_code=400)

        # Load treatment recommendations from treatment.json
        treatment_data = _load_treatment_data()
        
        # Get treatment for detected disease
        disease_key = disease  # e.g., "Corn___Healthy"
        treatment_info = treatment_data.get(disease_key, {}).get(lang, {})
        
        # Fallback to English if language not available
        if not treatment_info:
            treatment_info = treatment_data.get(disease_key, {}).get("en", {})
        
        treatment_title = treatment_info.get("title", disease.replace("___", " "))
        treatment_text = treatment_info.get("treatment", "Consult an agricultural extension officer for guidance.")
        
        # NEW: Estimate disease severity (affected percentage)
        affected_percentage = 0.0
        severity_level = 'low'
        
        if disease != "Corn___Healthy" and disease != "Corn___Not_Maize":
            # Only calculate severity for actual diseases
            affected_percentage, severity_level = model.estimate_severity(img)
        
        # Get detailed disease information with severity-specific recommendations
        disease_info = model.get_disease_info(disease)
        
        # Extract severity-specific message if available
        severity_modifiers = disease_info.get("severity_modifiers", {})
        severity_info = severity_modifiers.get(severity_level, {})
        
        # Build comprehensive recommendations
        recommendations = disease_info.get("management", [])
        prevention = disease_info.get("prevention", [])
        
        # Add severity-specific message at the top
        if severity_info.get("message"):
            recommendations.insert(0, severity_info["message"])
        
        # Determine if spray is recommended based on severity
        spray_recommended = severity_info.get("spray_recommended", True)
        
        # Check if we should show "Call AEO" button based on escalation rules
        show_call_aeo = False
        if severity_level == 'high' or confidence < 0.70:
            show_call_aeo = True
        
        # Save to Database
        status_text = "High Risk" if "Healthy" not in disease else "Healthy"
        crop = disease.split("___")[0]
        
        # Resolve real location from GPS coordinates
        location = "Unknown Region"
        if x_latitude and x_longitude:
            try:
                api_key = os.getenv("WEATHER_API_KEY", "")
                if api_key and api_key != "your_weather_api_key_here":
                    async with httpx.AsyncClient(timeout=3.0) as client:
                        resp = await client.get(
                            "https://api.openweathermap.org/data/2.5/weather",
                            params={"lat": x_latitude, "lon": x_longitude, "appid": api_key}
                        )
                        if resp.status_code == 200:
                            location = resp.json().get("name", "Unknown Region")
            except Exception as loc_err:
                print(f"[Location] Reverse lookup failed: {loc_err}")
        
        if device_id and not device_id.startswith("anonymous_"):
            database.register_farmer_scan(device_id, crop, disease, confidence, location, status_text, segment_id)
        
        response_data = {
            "disease": treatment_title,
            "disease_class": disease.replace("___", " "),
            "confidence": float(round(confidence, 2)),
            "treatment": treatment_text,
            "status": status_text,
            "recommendations": recommendations,
            "prevention": prevention,
            "disease_info": {
                "name": disease_info.get("name", treatment_title),
                "description": disease_info.get("description", ""),
                "symptoms": disease_info.get("symptoms", []),
                "scientific_name": disease_info.get("scientific_name", ""),
                "timing_window": disease_info.get("timing_window", "")
            },
            "severity": {
                "affected_percentage": float(round(affected_percentage, 1)),
                "level": severity_level,
                "message": severity_info.get("message", ""),
                "spray_recommended": spray_recommended,
                "range": severity_info.get("range", "")
            },
            "show_call_aeo": show_call_aeo,
            "location": location
        }
        
        print(f"[SUCCESS] Returning response: disease='{treatment_title}', confidence={confidence:.2f}")
        print(f"[SUCCESS] Full disease name: {disease}")
        
        return JSONResponse(response_data)
    
    except Exception as e:
        return JSONResponse({
            "error": f"Server error: {str(e)}",
            "disease": "System Error",
            "confidence": 0,
            "treatment": "",
            "recommendations": []
        }, status_code=500)


def _load_treatment_data() -> dict:
    """Load treatment data from treatment.json file"""
    try:
        with open("treatment.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"⚠ Could not load treatment.json: {e}")
        return {}

# ============================================================================
# WEATHER API (OpenWeatherMap - with caching & fallback)
# ============================================================================

# In-memory cache (60 second TTL)
_weather_cache: Dict[str, Any] = {}
_weather_cache_time: float = 0
WEATHER_CACHE_TTL = 60  # seconds

# Fallback data (used when API key is missing or call fails)
WEATHER_FALLBACK = {
    "temp": 32,
    "humidity": 85,
    "condition": "Partly Cloudy",
    "cloud": 60,
    "risk": "Medium",
    "city": "Kumasi",
    "source": "fallback"
}

def _calculate_risk(humidity: int) -> str:
    """Calculate disease outbreak risk level from humidity."""
    if humidity >= 80:
        return "High"
    elif humidity >= 60:
        return "Medium"
    return "Low"


@app.get("/weather")
async def get_weather(lat: Optional[float] = None, lon: Optional[float] = None):
    global _weather_cache, _weather_cache_time

    # Build a cache key based on location (rounded to 1 decimal for nearby reuse)
    if lat is not None and lon is not None:
        cache_key = f"{round(lat, 1)}_{round(lon, 1)}"
    else:
        cache_key = "default_city"

    # Return cached result if still fresh and same location
    now = time.time()
    if (_weather_cache 
        and _weather_cache.get("_cache_key") == cache_key 
        and (now - _weather_cache_time) < WEATHER_CACHE_TTL):
        return {k: v for k, v in _weather_cache.items() if not k.startswith("_")}

    # Read config
    api_key = os.getenv("WEATHER_API_KEY", "")
    api_url = os.getenv("WEATHER_API_URL", "https://api.openweathermap.org/data/2.5/weather")
    city = os.getenv("WEATHER_CITY", "Kumasi,GH")

    # No key set yet — return fallback silently
    if not api_key or api_key == "your_weather_api_key_here":
        return WEATHER_FALLBACK

    try:
        # Build request params — use coordinates if available, otherwise city name
        params = {"appid": api_key, "units": "metric"}
        if lat is not None and lon is not None:
            params["lat"] = lat
            params["lon"] = lon
        else:
            params["q"] = city

        async with httpx.AsyncClient(timeout=5.0) as client:
            resp = await client.get(api_url, params=params)
            resp.raise_for_status()
            raw = resp.json()

        temp = round(raw["main"]["temp"])
        humidity = raw["main"]["humidity"]
        condition = raw["weather"][0]["description"].title()
        city_name = raw["name"]
        cloud = raw.get("clouds", {}).get("all", 0)  # Cloud coverage percentage
        risk = _calculate_risk(humidity)

        result = {
            "temp": temp,
            "humidity": humidity,
            "condition": condition,
            "cloud": cloud,
            "risk": risk,
            "city": city_name,
            "source": "live"
        }

        # Cache result (with internal cache key)
        _weather_cache = {**result, "_cache_key": cache_key}
        _weather_cache_time = now
        return result

    except Exception as e:
        print(f"[Weather] API call failed: {e}. Using fallback.")
        return WEATHER_FALLBACK


# ============================================================================
# DASHBOARD & AUTH ROUTES
# ============================================================================

# Initialize DB on startup
@app.on_event("startup")
async def startup_event():
    print("🚀 [STARTUP] Initializing AgroGuard...")
    
    first_run = not os.path.exists("agroguard.db")
    
    # Always run init_db - it safely creates missing tables and columns
    print("📊 [STARTUP] Checking database schema...")
    database.init_db()
    
    if first_run:
        print("🆕 [STARTUP] First run detected - creating default accounts...")
        database.create_officer("admin", "admin123")  # Default officer credentials
        database.create_superadmin("superadmin", "SuperAdmin@123", "System Administrator")  # Default superadmin
        print("✅ [STARTUP] Database initialized with default officer and superadmin accounts")
    else:
        print("✅ [STARTUP] Database schema updated (if needed)")
        # Ensure superadmin exists even on subsequent runs (in case DB was created before superadmin feature)
        existing_superadmin = database.get_superadmin_by_username("superadmin")
        if not existing_superadmin:
            print("🔧 [STARTUP] Creating default superadmin account...")
            database.create_superadmin("superadmin", "SuperAdmin@123", "System Administrator")
            print("✅ [STARTUP] Default superadmin account created")
    
    print("✅ [STARTUP] AgroGuard backend ready!")



@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login")
async def login(
    request: Request,
    ghana_card: str = Form(None),
    staff_id: str = Form(None),
    password: str = Form(...)
):
    """
    AEO Officer login - accepts Ghana Card OR Staff ID with password created by SuperAdmin
    """
    # Determine which identifier was provided
    identifier = ghana_card or staff_id
    
    if not identifier:
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "Please provide either Ghana Card ID or Staff ID"
            }
        )
    
    # Retrieve AEO record by identifier
    aeo = database.get_aeo_by_identifier(identifier.strip())
    
    if not aeo:
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "Invalid credentials. Please check your ID and password."
            }
        )
    
    # Check if account is active
    if not aeo['is_active']:
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "Your account has been deactivated. Please contact administrator."
            }
        )
    
    # Verify password using passlib context from database module
    from passlib.context import CryptContext
    pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
    
    if not pwd_context.verify(password, aeo['hashed_password']):
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "Invalid credentials. Please check your ID and password."
            }
        )
    
    # Generate JWT token
    access_token = auth.create_access_token(
        data={
            "sub": str(aeo['id']),
            "type": "aeo",
            "staff_id": aeo['staff_id'],
            "name": aeo['name']
        }
    )
    
    # Check if profile is complete
    # Profile is complete if profile_completed flag is set to 1
    # sqlite3.Row can be accessed like a dict using keys() or direct indexing
    try:
        profile_complete = aeo['profile_completed'] == 1 if aeo['profile_completed'] is not None else False
    except (KeyError, TypeError):
        profile_complete = False
    
    # Update last_login timestamp
    conn = database.get_db_connection()
    conn.execute(
        "UPDATE aeo SET last_login = datetime('now') WHERE id = ?",
        (aeo['id'],)
    )
    conn.commit()
    conn.close()
    
    print(f"[LOGIN] AEO {aeo['staff_id']} - Profile complete: {profile_complete}")
    
    # Store AEO ID in session for profile updates
    request.session["aeo_id"] = aeo['id']
    request.session["aeo_name"] = aeo['name']
    request.session["aeo_staff_id"] = aeo['staff_id']
    
    # Set cookie
    response = RedirectResponse(
        url="/complete-profile" if not profile_complete else "/dashboard",
        status_code=status.HTTP_303_SEE_OTHER
    )
    response.set_cookie(key="access_token", value=access_token, httponly=True)
    return response


@app.get("/complete-profile", response_class=HTMLResponse)
async def complete_profile_page(request: Request, current_user: dict = Depends(auth.get_current_user)):
    """Profile completion page for first-time AEO login"""
    # Get current AEO data from database
    aeo_id = current_user["user_id"]
    conn = database.get_db_connection()
    cur = conn.cursor()
    aeo = cur.execute("SELECT * FROM aeo WHERE id = ?", (aeo_id,)).fetchone()
    conn.close()
    
    # If profile is already complete, redirect to dashboard
    if aeo and aeo['profile_completed'] == 1:
        return RedirectResponse(url="/dashboard", status_code=status.HTTP_303_SEE_OTHER)
    
    # Pass AEO data to template
    return templates.TemplateResponse("complete_profile.html", {
        "request": request,
        "aeo": dict(aeo) if aeo else {}
    })


@app.post("/api/aeo/complete-profile")
async def complete_profile(request: Request, current_user: dict = Depends(auth.get_current_user)):
    """Save AEO profile information and mark profile as completed"""
    try:
        data = await request.json()
        
        # Update AEO profile in database
        aeo_id = current_user["user_id"]
        
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            UPDATE aeo
            SET name = ?, email = ?, phone = ?, district = ?, profile_completed = 1
            WHERE id = ?
        """, (
            data.get('name'),
            data.get('email'),
            data.get('phone'),
            data.get('district'),
            aeo_id
        ))
        
        conn.commit()
        conn.close()
        
        print(f"[complete-profile] AEO {aeo_id} profile completed successfully")
        
        return JSONResponse({"success": True, "message": "Profile updated successfully"})
        
    except Exception as e:
        print(f"[complete-profile] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update profile: {str(e)}")


@app.post("/api/aeo/biometric/register")
async def register_biometric(request: Request, current_user: dict = Depends(auth.get_current_user)):
    """Register biometric (fingerprint) for AEO"""
    try:
        data = await request.json()
        aeo_id = current_user["user_id"]
        
        # Store biometric credentials
        # In a real implementation, this would verify and store the WebAuthn credential
        biometric_id = data.get('biometric_id')
        public_key = data.get('public_key')
        
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            UPDATE aeo
            SET biometric_id = ?, biometric_public_key = ?
            WHERE id = ?
        """, (biometric_id, public_key, aeo_id))
        
        conn.commit()
        conn.close()
        
        print(f"[biometric] AEO {aeo_id} biometric registered successfully")
        
        return JSONResponse({"success": True, "message": "Biometric registered successfully"})
        
    except Exception as e:
        print(f"[biometric] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to register biometric: {str(e)}")


@app.post("/api/aeo/upload-picture")
async def upload_profile_picture(
    file: UploadFile = File(...),
    current_user: dict = Depends(auth.get_current_user)
):
    """Upload AEO profile picture"""
    try:
        aeo_id = current_user["user_id"]
        
        # Validate file type
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="File must be an image")
        
        # Validate file size (max 5MB)
        contents = await file.read()
        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Image size must be less than 5MB")
        
        # Create uploads directory if it doesn't exist
        import os
        upload_dir = "static/uploads/profiles"
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate unique filename
        import uuid
        file_extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
        filename = f"aeo_{aeo_id}_{uuid.uuid4().hex[:8]}.{file_extension}"
        filepath = os.path.join(upload_dir, filename)
        
        # Save file
        with open(filepath, "wb") as f:
            f.write(contents)
        
        # Update database with profile picture path
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        # Delete old profile picture if exists
        old_picture = cur.execute(
            "SELECT profile_picture FROM aeo WHERE id = ?", (aeo_id,)
        ).fetchone()
        
        if old_picture and old_picture['profile_picture']:
            old_path = old_picture['profile_picture']
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except:
                    pass
        
        # Update with new picture path
        cur.execute(
            "UPDATE aeo SET profile_picture = ? WHERE id = ?",
            (filepath, aeo_id)
        )
        
        conn.commit()
        conn.close()
        
        print(f"[upload-picture] AEO {aeo_id} profile picture uploaded: {filename}")
        
        return JSONResponse({
            "success": True,
            "message": "Profile picture uploaded successfully",
            "picture_url": f"/{filepath}"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[upload-picture] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to upload picture: {str(e)}")


@app.post("/api/aeo/biometric/login")
async def biometric_login(request: Request):
    """Login with biometric (fingerprint)"""
    try:
        data = await request.json()
        biometric_id = data.get('biometric_id')
        
        if not biometric_id:
            raise HTTPException(status_code=400, detail="Biometric ID required")
        
        # Find AEO by biometric ID
        conn = database.get_db_connection()
        aeo = conn.execute(
            "SELECT * FROM aeo WHERE biometric_id = ? AND is_active = 1",
            (biometric_id,)
        ).fetchone()
        conn.close()
        
        if not aeo:
            raise HTTPException(status_code=401, detail="Biometric authentication failed")
        
        # Generate JWT token
        access_token = auth.create_access_token(
            data={
                "sub": str(aeo['id']),
                "type": "aeo",
                "staff_id": aeo['staff_id'],
                "name": aeo['name']
            }
        )
        
        # Update last_login
        conn = database.get_db_connection()
        conn.execute(
            "UPDATE aeo SET last_login = datetime('now') WHERE id = ?",
            (aeo['id'],)
        )
        conn.commit()
        conn.close()
        
        return JSONResponse({
            "success": True,
            "access_token": access_token,
            "token_type": "bearer"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[biometric-login] Error: {e}")
        raise HTTPException(status_code=500, detail="Biometric login failed")


@app.get("/signup", response_class=HTMLResponse)
async def signup_page(request: Request):
    """Signup disabled - AEO accounts are created by SuperAdmin only"""
    return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/signup")
async def signup(request: Request):
    """Signup disabled - AEO accounts are created by SuperAdmin only"""
    return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/api/aeo/login", response_model=schemas.TokenResponse)
async def aeo_login_api(login_data: schemas.AEOLoginRequest):
    """
    AEO login API endpoint (for mobile/API clients) - accepts staff_id, ghana_card, or phone as identifier
    along with password. Returns JWT token on success.
    """
    # Determine which identifier was provided
    identifier = login_data.staff_id or login_data.ghana_card or login_data.phone
    
    if not identifier:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Must provide at least one identifier (staff_id, ghana_card, or phone)"
        )
    
    if not login_data.password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Password is required"
        )
    
    # Retrieve AEO record
    aeo = database.get_aeo_by_identifier(identifier)
    
    if not aeo:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials"
        )
    
    # Check if account is active
    if not aeo['is_active']:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account has been deactivated"
        )
    
    # Verify password using passlib context from database module
    from passlib.context import CryptContext
    pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
    
    if not pwd_context.verify(login_data.password, aeo['hashed_password']):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials"
        )
    
    # Generate JWT token
    access_token = auth.create_access_token(
        data={
            "sub": str(aeo['id']),
            "type": "aeo",
            "staff_id": aeo['staff_id']
        }
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user_id": aeo['id'],
        "user_type": "aeo"
    }


@app.get("/superadmin/login", response_class=HTMLResponse)
async def superadmin_login_page(request: Request):
    """Super Admin login page"""
    return templates.TemplateResponse("superadmin_login.html", {"request": request, "error": None})


@app.post("/superadmin/login")
async def superadmin_login_form(request: Request, username: str = Form(...), password: str = Form(...)):
    """Super Admin login form handler (for HTML form submission with redirect)"""
    superadmin = database.verify_superadmin(username, password)
    if not superadmin:
        return templates.TemplateResponse("superadmin_login.html", {"request": request, "error": "Invalid credentials"})
    
    # Generate JWT
    access_token = auth.create_access_token(
        data={
            "sub": str(superadmin['id']),
            "type": "superadmin",
            "username": superadmin['username']
        }
    )
    
    response = RedirectResponse(url="/superadmin/dashboard", status_code=status.HTTP_303_SEE_OTHER)
    response.set_cookie(key="access_token", value=access_token, httponly=True)
    return response


@app.get("/superadmin/dashboard", response_class=HTMLResponse)
async def superadmin_dashboard(request: Request, current_user: dict = Depends(auth.get_superadmin_user)):
    """Super Admin dashboard page"""
    return templates.TemplateResponse("superadmin_dashboard.html", {
        "request": request,
        "user": current_user.get("username", "Super Admin")
    })


@app.post("/api/superadmin/login", response_model=schemas.TokenResponse)
async def superadmin_api_login(username: str = Form(...), password: str = Form(...)):
    """
    Super Admin API login endpoint - accepts username and password.
    Returns JWT token with superadmin privileges on success (for API clients).
    """
    if not username or not password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username and password are required"
        )
    
    # Verify super admin credentials
    superadmin = database.verify_superadmin(username, password)
    
    if not superadmin:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials"
        )
    
    # Generate JWT token with superadmin type
    access_token = auth.create_access_token(
        data={
            "sub": str(superadmin['id']),
            "type": "superadmin",
            "username": superadmin['username']
        }
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user_id": superadmin['id'],
        "user_type": "superadmin"
    }


@app.post("/superadmin/aeo/create", response_model=schemas.AEOResponse)
async def create_aeo_account(
    aeo_data: schemas.AEOCreateRequest,
    current_user: dict = Depends(auth.get_superadmin_user)
):
    """
    Super Admin only endpoint to create new AEO accounts.
    Requires superadmin JWT token. Creates AEO with temporary password
    that must be changed on first login.
    """
    try:
        # Create the AEO account
        aeo_id = database.create_aeo(
            staff_id=aeo_data.staff_id,
            ghana_card=aeo_data.ghana_card,
            phone=aeo_data.phone,
            name=aeo_data.name,
            password=aeo_data.temporary_password
        )
        
        if not aeo_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="AEO account creation failed. Staff ID, Ghana Card, or Phone may already exist."
            )
        
        # Log the action in audit log
        database.log_audit(
            action="create_aeo",
            entity="aeo",
            entity_id=aeo_id,
            performed_by=current_user["user_id"],
            details=f"Created AEO account for {aeo_data.name} (Staff ID: {aeo_data.staff_id})"
        )
        
        # Retrieve the created AEO to return
        aeo = database.get_aeo_by_identifier(aeo_data.staff_id)
        
        return {
            "id": aeo['id'],
            "staff_id": aeo['staff_id'],
            "ghana_card": aeo['ghana_card'],
            "phone": aeo['phone'],
            "name": aeo['name'],
            "must_change_password": bool(aeo['must_change_password']),
            "is_active": bool(aeo['is_active'])
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create AEO account: {str(e)}"
        )


@app.get("/api/superadmin/aeo/list")
async def list_all_aeos(current_user: dict = Depends(auth.get_superadmin_user)):
    """
    Super Admin endpoint to list all AEO accounts with their details and statistics
    """
    try:
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        # Get all AEOs
        aeos = cur.execute("""
            SELECT id, staff_id, ghana_card, phone, name, email, district,
                   must_change_password, is_active
            FROM aeo
            ORDER BY id DESC
        """).fetchall()
        
        # Get stats
        total_count = len(aeos)
        active_count = sum(1 for aeo in aeos if aeo['is_active'])
        
        # Get recent audit log count (last 7 days)
        recent_actions = cur.execute("""
            SELECT COUNT(*) as count
            FROM audit_log
            WHERE performed_by = ? AND timestamp >= datetime('now', '-7 days')
        """, (current_user["user_id"],)).fetchone()['count']
        
        conn.close()
        
        return {
            "aeos": [dict(aeo) for aeo in aeos],
            "stats": {
                "total": total_count,
                "active": active_count,
                "inactive": total_count - active_count,
                "recent_actions": recent_actions
            }
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch AEO list: {str(e)}"
        )


@app.put("/api/superadmin/aeo/{aeo_id}/toggle")
async def toggle_aeo_status(
    aeo_id: int,
    current_user: dict = Depends(auth.get_superadmin_user)
):
    """
    Super Admin endpoint to activate/deactivate an AEO account
    """
    try:
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        # Get current status
        aeo = cur.execute("SELECT * FROM aeo WHERE id = ?", (aeo_id,)).fetchone()
        
        if not aeo:
            conn.close()
            raise HTTPException(status_code=404, detail="AEO not found")
        
        # Toggle status
        new_status = 0 if aeo['is_active'] else 1
        cur.execute("UPDATE aeo SET is_active = ? WHERE id = ?", (new_status, aeo_id))
        conn.commit()
        
        # Log the action
        database.log_audit(
            action="toggle_aeo_status",
            entity="aeo",
            entity_id=aeo_id,
            performed_by=current_user["user_id"],
            details=f"{'Activated' if new_status else 'Deactivated'} AEO account: {aeo['name']} ({aeo['staff_id']})"
        )
        
        conn.close()
        
        return {
            "success": True,
            "message": f"AEO account {'activated' if new_status else 'deactivated'} successfully",
            "is_active": bool(new_status)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to toggle AEO status: {str(e)}"
        )


@app.delete("/api/superadmin/aeo/{aeo_id}")
async def delete_aeo_account(
    aeo_id: int,
    current_user: dict = Depends(auth.get_superadmin_user)
):
    """
    Super Admin endpoint to permanently delete an AEO account
    WARNING: This cannot be undone!
    """
    try:
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        # Get AEO details before deletion
        aeo = cur.execute("SELECT * FROM aeo WHERE id = ?", (aeo_id,)).fetchone()
        
        if not aeo:
            conn.close()
            raise HTTPException(status_code=404, detail="AEO not found")
        
        # Log the action BEFORE deleting
        database.log_audit(
            action="delete_aeo",
            entity="aeo",
            entity_id=aeo_id,
            performed_by=current_user["user_id"],
            details=f"Deleted AEO account: {aeo['name']} ({aeo['staff_id']}) - Ghana Card: {aeo['ghana_card']}"
        )
        
        # Delete the AEO
        cur.execute("DELETE FROM aeo WHERE id = ?", (aeo_id,))
        conn.commit()
        conn.close()
        
        return {
            "success": True,
            "message": f"AEO account for {aeo['name']} deleted permanently"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete AEO: {str(e)}"
        )


@app.get("/api/superadmin/audit-log")
async def get_audit_log(
    limit: int = Query(50, le=200),
    current_user: dict = Depends(auth.get_superadmin_user)
):
    """
    Super Admin endpoint to view audit log of all admin actions
    """
    try:
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        logs = cur.execute("""
            SELECT 
                al.id,
                al.action,
                al.entity,
                al.entity_id,
                al.timestamp,
                al.details,
                sa.username as performed_by_username
            FROM audit_log al
            LEFT JOIN superadmin sa ON al.performed_by = sa.id
            ORDER BY al.timestamp DESC
            LIMIT ?
        """, (limit,)).fetchall()
        
        conn.close()
        
        return {
            "logs": [dict(log) for log in logs],
            "total": len(logs)
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch audit log: {str(e)}"
        )


# Removed duplicate login route



@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    response.delete_cookie("access_token")
    return response


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, current_user: dict = Depends(auth.get_current_user)):
    # Get AEO info including profile picture
    conn = database.get_db_connection()
    cur = conn.cursor()
    aeo = cur.execute(
        "SELECT name, district, profile_picture, phone, email, region FROM aeo WHERE id = ?",
        (current_user["user_id"],)
    ).fetchone()
    conn.close()
    
    # Store in session for profile updates
    request.session["aeo_id"] = current_user["user_id"]
    if aeo:
        request.session["aeo_name"] = aeo['name']
    
    aeo_info = {
        "name": aeo['name'] if aeo else current_user["user_id"],
        "district": aeo['district'] if aeo and aeo['district'] else "N/A",
        "profile_picture": aeo['profile_picture'] if aeo and aeo['profile_picture'] else None,
        "phone": aeo['phone'] if aeo and aeo['phone'] else "",
        "email": aeo['email'] if aeo and aeo['email'] else "",
        "region": aeo['region'] if aeo and aeo['region'] else ""
    }
    
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": aeo_info["name"],
        "district": aeo_info["district"],
        "profile_picture": aeo_info["profile_picture"],
        "phone": aeo_info["phone"],
        "email": aeo_info["email"],
        "region": aeo_info["region"]
    })


@app.get("/api/dashboard/stats")
async def dashboard_stats(current_user: dict = Depends(auth.get_current_user)):
    stats = database.get_dashboard_stats()
    return JSONResponse(stats)


@app.post("/register-farmer")
async def register_farmer_profile(request: Request):
    """
    Called once from the farmer app on first launch.
    Saves the farmer's real name and phone number linked to their device ID.
    """
    try:
        data = await request.json()
        device_id = data.get("device_id", "").strip()
        name      = data.get("name", "").strip()
        phone     = data.get("phone", "").strip()

        if not device_id or not name or not phone:
            raise HTTPException(status_code=400, detail="device_id, name and phone are required.")

        database.register_farmer_profile(device_id, name, phone)
        return JSONResponse({"success": True, "message": "Farmer profile saved."})

    except HTTPException:
        raise
    except Exception as e:
        print(f"[register-farmer] Error: {e}")

# -------------------------------------------------
# NEW ENDPOINT – List all registered farmers for the dashboard
# -------------------------------------------------
@app.get("/api/farmers", response_class=JSONResponse)
async def get_farmers(current_user: dict = Depends(auth.get_current_user)):
    """
    Return a list of all farmers with their device_id, name, phone and last_seen.
    Only accessible to authenticated officers.
    """
    conn = database.get_db_connection()
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT device_id, name, phone, last_seen
        FROM farmers
        ORDER BY last_seen DESC
        """
    ).fetchall()
    conn.close()
    farmers = [
        {
            "device_id": r["device_id"],
            "name": r["name"] or "Anonymous",
            "phone": r["phone"] or "—",
            "last_seen": r["last_seen"],
        }
        for r in rows
    ]
    return JSONResponse({"farmers": farmers})


# ============================================================================
# NEW FEATURES: Add Farmer, Send Alert, Support Ticket
# ============================================================================

@app.post("/api/farmer/add")
async def add_farmer_manually(request: Request, current_user: dict = Depends(auth.get_current_user)):
    """
    AEO manually adds a farmer (for farmers without smartphones).
    Creates a virtual device_id and stores farmer information.
    """
    try:
        data = await request.json()
        
        # Validate required fields
        name = data.get('name', '').strip()
        phone = data.get('phone', '').strip()
        district = data.get('district', '').strip()
        
        if not name or not phone or not district:
            raise HTTPException(
                status_code=400,
                detail="Name, phone, and district are required"
            )
        
        # Optional fields
        ghana_card = data.get('ghana_card', '').strip() or None
        crops = data.get('crops', '').strip() or None
        
        # Generate a virtual device_id for manually added farmers
        import hashlib
        device_id = f"manual_{hashlib.md5(phone.encode()).hexdigest()[:16]}"
        
        # Check if farmer already exists
        conn = database.get_db_connection()
        cur = conn.cursor()
        existing = cur.execute(
            "SELECT id FROM farmers WHERE phone = ? OR device_id = ?",
            (phone, device_id)
        ).fetchone()
        
        if existing:
            conn.close()
            raise HTTPException(
                status_code=400,
                detail="A farmer with this phone number already exists"
            )
        
        # Insert farmer
        cur.execute("""
            INSERT INTO farmers (device_id, name, phone, ghana_card, district, crops, registration_method, registered_by, last_seen)
            VALUES (?, ?, ?, ?, ?, ?, 'manual', ?, datetime('now'))
        """, (device_id, name, phone, ghana_card, district, crops, current_user['user_id']))
        
        farmer_id = cur.lastrowid
        conn.commit()
        conn.close()
        
        # Log the action
        database.log_audit(
            action="add_farmer_manual",
            entity="farmer",
            entity_id=farmer_id,
            performed_by=current_user['user_id'],
            details=f"Manually added farmer: {name} ({phone})"
        )
        
        return JSONResponse({
            "success": True,
            "message": f"Farmer '{name}' added successfully",
            "farmer_id": farmer_id,
            "device_id": device_id
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[add-farmer] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to add farmer: {str(e)}")


@app.post("/api/alert/send")
async def send_alert_to_farmers(request: Request, current_user: dict = Depends(auth.get_current_user)):
    """
    AEO sends broadcast alert/notification to farmers.
    Supports: all farmers, region, district, crop type, or specific phone numbers.
    """
    try:
        data = await request.json()
        
        # Validate required fields - handle None values properly
        alert_type = (data.get('alert_type') or '').strip()
        target_type = (data.get('target_type') or '').strip()
        priority = (data.get('priority') or '').strip()
        title = (data.get('title') or '').strip()
        message = (data.get('message') or '').strip()
        
        if not alert_type or not target_type or not priority or not title or not message:
            raise HTTPException(
                status_code=400,
                detail="Alert type, target type, priority, title, and message are required"
            )
        
        # Optional targeting fields - handle None values
        region = (data.get('region') or '').strip() or None
        district = (data.get('district') or '').strip() or None
        crop = (data.get('crop') or '').strip() or None
        phone_numbers = (data.get('phone_numbers') or '').strip() or None
        
        # Build query to find target farmers
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        if target_type == 'all':
            # All farmers with phone numbers
            query = "SELECT device_id, name, phone FROM farmers WHERE phone IS NOT NULL AND phone != ''"
            params = ()
            target_description = "All Farmers (Broadcast)"
            
        elif target_type == 'region':
            # Farmers in specific region (and optionally specific district)
            if not region:
                raise HTTPException(status_code=400, detail="Region must be specified")
            
            if district:
                # Specific district within region
                query = "SELECT device_id, name, phone FROM farmers WHERE district = ? AND phone IS NOT NULL AND phone != ''"
                params = (district,)
                target_description = f"{district}, {region}"
            else:
                # All districts in region (would need region-district mapping)
                # For now, treat as district search
                query = "SELECT device_id, name, phone FROM farmers WHERE phone IS NOT NULL AND phone != ''"
                params = ()
                target_description = f"{region} Region"
                
        elif target_type == 'district':
            # Farmers in specific district
            if not district:
                raise HTTPException(status_code=400, detail="District must be specified")
            
            query = "SELECT device_id, name, phone FROM farmers WHERE district = ? AND phone IS NOT NULL AND phone != ''"
            params = (district,)
            target_description = f"{district} District"
            
        elif target_type == 'crop':
            # Farmers growing specific crop
            if not crop:
                raise HTTPException(status_code=400, detail="Crop type must be specified")
            
            query = "SELECT device_id, name, phone FROM farmers WHERE crops LIKE ? AND phone IS NOT NULL AND phone != ''"
            params = (f"%{crop}%",)
            target_description = f"Farmers growing {crop}"
            
        elif target_type == 'phone':
            # Specific phone number(s)
            if not phone_numbers:
                raise HTTPException(status_code=400, detail="Phone numbers must be provided")
            
            # Parse comma-separated phone numbers
            phones = [p.strip() for p in phone_numbers.split(',') if p.strip()]
            if not phones:
                raise HTTPException(status_code=400, detail="No valid phone numbers provided")
            
            # Build query with phone numbers
            placeholders = ','.join(['?' for _ in phones])
            query = f"SELECT device_id, name, phone FROM farmers WHERE phone IN ({placeholders})"
            params = tuple(phones)
            target_description = f"{len(phones)} specific farmer(s)"
            
        else:
            raise HTTPException(status_code=400, detail="Invalid target type")
        
        # Execute query
        farmers = cur.execute(query, params).fetchall()
        recipient_count = len(farmers)
        
        if recipient_count == 0:
            conn.close()
            raise HTTPException(
                status_code=400,
                detail=f"No farmers found for target: {target_description}"
            )
        
        # Store alert in database
        cur.execute("""
            INSERT INTO alerts (
                alert_type, title, message, priority,
                target_audience, target_type, target_phone, district,
                sent_by, recipient_count, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
        """, (
            alert_type, title, message, priority,
            target_description,  # target_audience (for backward compatibility)
            target_description,  # target_type (new field)
            phone_numbers if target_type == 'phone' else None,
            district,
            current_user['user_id'],
            recipient_count
        ))
        
        alert_id = cur.lastrowid
        conn.commit()
        conn.close()
        
        # Log the action
        database.log_audit(
            action="send_alert",
            entity="alert",
            entity_id=alert_id,
            performed_by=current_user['user_id'],
            details=f"Sent {priority} priority {alert_type} alert to {recipient_count} farmers: {target_description}"
        )
        
        # TODO: In production, integrate with SMS gateway (Twilio, Africa's Talking, etc.)
        # For now, alerts are stored in database and accessible via farmer app
        
        return JSONResponse({
            "success": True,
            "message": f"Alert sent successfully to {target_description}",
            "recipient_count": recipient_count,
            "alert_id": alert_id
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[send-alert] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send alert: {str(e)}")


@app.get("/api/alert/estimate-recipients")
async def estimate_alert_recipients(
    target_type: str,
    region: str = "",
    district: str = "",
    crop: str = "",
    current_user: dict = Depends(auth.get_current_user)
):
    """Estimate number of recipients for an alert based on targeting"""
    try:
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        if target_type == 'all':
            count = cur.execute(
                "SELECT COUNT(*) as cnt FROM farmers WHERE phone IS NOT NULL AND phone != ''"
            ).fetchone()['cnt']
            
        elif target_type == 'region' or target_type == 'district':
            if district:
                count = cur.execute(
                    "SELECT COUNT(*) as cnt FROM farmers WHERE district = ? AND phone IS NOT NULL AND phone != ''",
                    (district,)
                ).fetchone()['cnt']
            else:
                count = cur.execute(
                    "SELECT COUNT(*) as cnt FROM farmers WHERE phone IS NOT NULL AND phone != ''"
                ).fetchone()['cnt']
                
        elif target_type == 'crop':
            if crop:
                count = cur.execute(
                    "SELECT COUNT(*) as cnt FROM farmers WHERE crops LIKE ? AND phone IS NOT NULL AND phone != ''",
                    (f"%{crop}%",)
                ).fetchone()['cnt']
            else:
                count = 0
                
        elif target_type == 'phone':
            count = 0  # Will depend on how many phone numbers are entered
            
        else:
            count = 0
        
        conn.close()
        
        return JSONResponse({"count": count})
        
    except Exception as e:
        print(f"[estimate-recipients] Error: {e}")
        return JSONResponse({"count": 0})
        if target == 'district':
            recipients_text += f" in {district}"
        
        return JSONResponse({
            "success": True,
            "message": "Alert sent successfully",
            "alert_id": alert_id,
            "recipients": recipients_text,
            "recipient_count": recipient_count
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[send-alert] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send alert: {str(e)}")


@app.post("/api/support/ticket")
async def create_support_ticket(request: Request, current_user: dict = Depends(auth.get_current_user)):
    """
    AEO submits a support ticket for technical issues or help requests.
    Creates a ticket that can be tracked and resolved by system administrators.
    """
    try:
        data = await request.json()
        
        # Validate required fields
        category = data.get('category', '').strip()
        priority = data.get('priority', '').strip()
        subject = data.get('subject', '').strip()
        description = data.get('description', '').strip()
        
        if not category or not priority or not subject or not description:
            raise HTTPException(
                status_code=400,
                detail="Category, priority, subject, and description are required"
            )
        
        # Optional fields
        contact = data.get('contact', '').strip() or None
        
        # Insert ticket into database
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            INSERT INTO support_tickets (
                category, priority, subject, description, contact, 
                submitted_by, status, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, 'open', datetime('now'))
        """, (category, priority, subject, description, contact, current_user['user_id']))
        
        ticket_id = cur.lastrowid
        conn.commit()
        conn.close()
        
        # Log the action
        database.log_audit(
            action="create_support_ticket",
            entity="support_ticket",
            entity_id=ticket_id,
            performed_by=current_user['user_id'],
            details=f"Created {priority} priority ticket: {subject}"
        )
        
        return JSONResponse({
            "success": True,
            "message": "Support ticket submitted successfully",
            "ticket_id": f"TICK-{ticket_id:05d}"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[support-ticket] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to create ticket: {str(e)}")


# ============================================================================
# FARMER NOTIFICATIONS & ALERTS API
# ============================================================================

@app.get("/api/farmer/notifications")
async def get_farmer_notifications(request: Request):
    """
    Get general notifications for farmer (app upgrades, weather notices, tips).
    """
    try:
        device_id = request.headers.get('device-id', '')
        
        # Generate timestamps dynamically based on current date
        now = datetime.now()
        
        # For now, return mock notifications with current dates
        # In production, these would come from a notifications table
        notifications = [
            {
                "id": "notif_1",
                "type": "info",
                "title": "Welcome to AgroGuard!",
                "message": "Get instant maize disease detection and connect with agricultural extension officers for expert support.",
                "timestamp": (now - timedelta(days=2)).isoformat(),  # 2 days ago
                "read": True
            },
            {
                "id": "notif_2",
                "type": "warning",
                "title": "Weather Alert",
                "message": "Heavy rains expected in your area over the next 3 days. High humidity may increase disease risk. Monitor your crops closely.",
                "timestamp": (now - timedelta(hours=6)).isoformat(),  # 6 hours ago
                "read": False
            },
            {
                "id": "notif_3",
                "type": "success",
                "title": "Farming Tip",
                "message": "Did you know? Proper spacing between maize plants (75cm apart) improves air circulation and reduces disease spread.",
                "timestamp": (now - timedelta(minutes=30)).isoformat(),  # 30 minutes ago
                "read": False
            }
        ]
        
        return {"notifications": notifications}
        
    except Exception as e:
        print(f"[ERROR] Failed to load notifications: {e}")
        return {"notifications": []}


@app.get("/api/farmer/alerts")
async def get_farmer_alerts(request: Request):
    """
    Get alerts sent by AEO dashboard to this farmer.
    Connects to the AEO alert system.
    """
    try:
        device_id = request.headers.get('device-id', '')
        phone = request.headers.get('phone', '')
        
        if not phone:
            return {"alerts": []}
        
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        # Get farmer information to match alerts
        farmer = cur.execute("""
            SELECT district, crops FROM farmers 
            WHERE phone = ? OR device_id = ?
            LIMIT 1
        """, (phone, device_id)).fetchone()
        
        if not farmer:
            conn.close()
            return {"alerts": []}
        
        farmer_district = farmer['district'] if farmer else None
        farmer_crops = farmer['crops'] if farmer else ''
        
        # Fetch alerts from the alerts table (sent by AEOs)
        # Match based on: all farmers, phone, district, or crop
        alerts_data = cur.execute("""
            SELECT id, title, message, priority, sent_by, created_at, target_type, district
            FROM alerts
            WHERE 
                target_type LIKE '%All Farmers%'
                OR target_type LIKE '%Broadcast%'
                OR target_phone LIKE '%' || ? || '%'
                OR (district = ? AND district IS NOT NULL)
                OR (target_type LIKE '%' || ? || '%')
            ORDER BY created_at DESC
            LIMIT 20
        """, (phone, farmer_district, farmer_district)).fetchall()
        
        conn.close()
        
        alerts = []
        for row in alerts_data:
            # Get AEO name
            conn2 = database.get_db_connection()
            aeo = conn2.execute(
                "SELECT name FROM aeo WHERE id = ?", (row['sent_by'],)
            ).fetchone()
            conn2.close()
            
            alerts.append({
                "id": str(row['id']),
                "title": row['title'],
                "message": row['message'],
                "priority": row['priority'] or 'medium',
                "from": aeo['name'] if aeo else 'Extension Officer',
                "timestamp": row['created_at'],
                "read": False  # Can implement read tracking later
            })
        
        return {"alerts": alerts}
        
    except Exception as e:
        print(f"[ERROR] Failed to load alerts: {e}")
        import traceback
        traceback.print_exc()
        return {"alerts": []}


@app.post("/api/farmer/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, request: Request):
    """
    Mark a notification as read.
    """
    # In production, update read status in database
    return {"success": True}


@app.post("/api/farmer/alerts/{alert_id}/read")
async def mark_alert_read(alert_id: str, request: Request):
    """
    Mark an alert as read.
    """
    try:
        conn = database.get_db_connection()
        cur = conn.cursor()
        
        # Update alert read status (if we add read tracking)
        # For now, just return success
        
        conn.close()
        return {"success": True}
        
    except Exception as e:
        print(f"[ERROR] Failed to mark alert as read: {e}")
        return {"success": False}


# ============================================================================
# EXPORT DATA ENDPOINTS
# ============================================================================

@app.get("/api/export/farmers")
async def export_farmers(format: str = "csv", current_user: dict = Depends(auth.get_current_user)):
    """Export farmers data to CSV or Excel"""
    try:
        import io
        import csv
        
        conn = database.get_db_connection()
        farmers = conn.execute("""
            SELECT device_id, name, phone, ghana_card, district, crops, 
                   registration_method, first_seen, last_seen
            FROM farmers
            ORDER BY last_seen DESC
        """).fetchall()
        conn.close()
        
        if format.lower() == "csv":
            # Create CSV
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow(['Device ID', 'Name', 'Phone', 'Ghana Card', 'District', 
                           'Crops', 'Registration Method', 'First Seen', 'Last Seen'])
            
            # Write data
            for farmer in farmers:
                writer.writerow([
                    farmer['device_id'],
                    farmer['name'] or 'N/A',
                    farmer['phone'] or 'N/A',
                    farmer['ghana_card'] or 'N/A',
                    farmer['district'] or 'N/A',
                    farmer['crops'] or 'N/A',
                    farmer['registration_method'] or 'app',
                    farmer['first_seen'],
                    farmer['last_seen']
                ])
            
            # Return CSV
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=farmers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
            )
        
        elif format.lower() == "excel":
            # Create Excel using openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Farmers"
                
                # Header styling
                header_fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                # Write header
                headers = ['Device ID', 'Name', 'Phone', 'Ghana Card', 'District', 
                          'Crops', 'Registration Method', 'First Seen', 'Last Seen']
                ws.append(headers)
                
                # Style header
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data
                for farmer in farmers:
                    ws.append([
                        farmer['device_id'],
                        farmer['name'] or 'N/A',
                        farmer['phone'] or 'N/A',
                        farmer['ghana_card'] or 'N/A',
                        farmer['district'] or 'N/A',
                        farmer['crops'] or 'N/A',
                        farmer['registration_method'] or 'app',
                        farmer['first_seen'],
                        farmer['last_seen']
                    ])
                
                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Save to bytes
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                return Response(
                    content=output.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=farmers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
                )
                
            except ImportError:
                raise HTTPException(status_code=500, detail="Excel export requires openpyxl library. Install with: pip install openpyxl")
        
        else:
            raise HTTPException(status_code=400, detail="Format must be 'csv' or 'excel'")
            
    except Exception as e:
        print(f"[export-farmers] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@app.get("/api/export/scans")
async def export_scans(format: str = "csv", current_user: dict = Depends(auth.get_current_user)):
    """Export scans data to CSV or Excel"""
    try:
        import io
        import csv
        
        conn = database.get_db_connection()
        scans = conn.execute("""
            SELECT s.id, s.farmer_device_id, f.name as farmer_name, f.phone,
                   s.crop, s.disease, s.confidence, s.location, s.status, s.timestamp
            FROM scans s
            LEFT JOIN farmers f ON s.farmer_device_id = f.device_id
            ORDER BY s.timestamp DESC
        """).fetchall()
        conn.close()
        
        if format.lower() == "csv":
            # Create CSV
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow(['Scan ID', 'Device ID', 'Farmer Name', 'Phone', 'Crop', 
                           'Disease', 'Confidence', 'Location', 'Status', 'Timestamp'])
            
            # Write data
            for scan in scans:
                writer.writerow([
                    scan['id'],
                    scan['farmer_device_id'],
                    scan['farmer_name'] or 'N/A',
                    scan['phone'] or 'N/A',
                    scan['crop'] or 'N/A',
                    scan['disease'],
                    f"{scan['confidence']:.2%}" if scan['confidence'] else 'N/A',
                    scan['location'] or 'N/A',
                    scan['status'],
                    scan['timestamp']
                ])
            
            # Return CSV
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=scans_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
            )
        
        elif format.lower() == "excel":
            # Create Excel
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Disease Scans"
                
                # Header styling
                header_fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                # Write header
                headers = ['Scan ID', 'Device ID', 'Farmer Name', 'Phone', 'Crop', 
                          'Disease', 'Confidence', 'Location', 'Status', 'Timestamp']
                ws.append(headers)
                
                # Style header
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data
                for scan in scans:
                    ws.append([
                        scan['id'],
                        scan['farmer_device_id'],
                        scan['farmer_name'] or 'N/A',
                        scan['phone'] or 'N/A',
                        scan['crop'] or 'N/A',
                        scan['disease'],
                        f"{scan['confidence']:.2%}" if scan['confidence'] else 'N/A',
                        scan['location'] or 'N/A',
                        scan['status'],
                        scan['timestamp']
                    ])
                
                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Save to bytes
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                return Response(
                    content=output.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=scans_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
                )
                
            except ImportError:
                raise HTTPException(status_code=500, detail="Excel export requires openpyxl library. Install with: pip install openpyxl")
        
        else:
            raise HTTPException(status_code=400, detail="Format must be 'csv' or 'excel'")
            
    except Exception as e:
        print(f"[export-scans] Error: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# ============================================================================
# FARMER API ENDPOINTS - Call AEO & Support
# ============================================================================

@app.get("/api/farmer/get-all-aeos")
async def get_all_available_aeos(request: Request):
    """
    Get all active AEOs in the system for farmer to contact.
    Returns list of all available Extension Officers with their contact info.
    """
    try:
        # Query all active AEOs
        conn = database.get_db_connection()
        c = conn.cursor()
        
        aeos = c.execute('''
            SELECT id, name, phone, district, region, email, profile_picture
            FROM aeo
            WHERE is_active = 1
            ORDER BY name ASC
        ''').fetchall()
        
        conn.close()
        
        if aeos:
            return {
                "success": True,
                "count": len(aeos),
                "aeos": [
                    {
                        "id": aeo['id'],
                        "name": aeo['name'],
                        "phone": aeo['phone'],
                        "district": aeo['district'] if aeo['district'] else "N/A",
                        "region": aeo['region'] if aeo['region'] else "N/A",
                        "email": aeo['email'] if aeo['email'] else None,
                        "profile_picture": aeo['profile_picture'] if aeo['profile_picture'] else None
                    }
                    for aeo in aeos
                ]
            }
        else:
            return {
                "success": True,
                "count": 0,
                "aeos": [],
                "message": "No Extension Officers currently registered in the system"
            }
            
    except Exception as e:
        print(f"[API] Error fetching all AEOs: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch AEO list: {str(e)}")


@app.get("/api/farmer/get-district-from-gps")
async def get_district_from_gps(
    lat: float = Query(..., description="Latitude"),
    lon: float = Query(..., description="Longitude")
):
    """
    Convert GPS coordinates to Ghana district using reverse geocoding.
    Uses OpenWeatherMap's geocoding API or fallback to simple district mapping.
    """
    try:
        # Simple Ghana district boundaries (approximate center points)
        # In production, use a proper geocoding API or shapefile lookup
        ghana_districts = {
            "Greater Accra": {
                "Accra Metro": (5.6037, -0.1870),
                "Tema": (5.6698, -0.0166),
                "Ga East": (5.7333, -0.2167),
                "Ga West": (5.6500, -0.3500),
            },
            "Ashanti": {
                "Kumasi": (6.6885, -1.6244),
                "Obuasi": (6.2027, -1.6708),
                "Ejisu": (6.6500, -1.4833),
            },
            "Western": {
                "Sekondi-Takoradi": (4.9344, -1.7817),
                "Tarkwa-Nsuaem": (5.2994, -1.9981),
            },
            "Eastern": {
                "Koforidua": (6.0939, -0.2592),
                "New Juaben": (6.0833, -0.2667),
            },
            "Central": {
                "Cape Coast": (5.1053, -1.2466),
                "Komenda-Edina-Eguafo-Abirem": (5.0833, -1.4167),
            },
            "Northern": {
                "Tamale": (9.4034, -0.8424),
                "Yendi": (9.4439, -0.0103),
            },
            "Upper East": {
                "Bolgatanga": (10.7856, -0.8514),
                "Bawku": (11.0619, -0.2419),
            },
            "Upper West": {
                "Wa": (10.0608, -2.5097),
            },
        }
        
        # Find nearest district by calculating distance
        min_distance = float('inf')
        nearest_district = None
        
        for region, districts in ghana_districts.items():
            for district, (d_lat, d_lon) in districts.items():
                # Simple Euclidean distance (good enough for district-level matching)
                distance = ((lat - d_lat) ** 2 + (lon - d_lon) ** 2) ** 0.5
                if distance < min_distance:
                    min_distance = distance
                    nearest_district = district
        
        if nearest_district:
            return {
                "success": True,
                "district": nearest_district,
                "lat": lat,
                "lon": lon,
                "method": "proximity_matching"
            }
        else:
            raise HTTPException(status_code=404, detail="Could not determine district from coordinates")
            
    except Exception as e:
        print(f"[API] Error determining district from GPS: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to determine district: {str(e)}")


@app.get("/api/farmer/get-aeo")
async def get_assigned_aeo(
    request: Request,
    district: str = Query(..., description="Farmer's district")
):
    """
    Get AEO assigned to a farmer's district.
    Returns AEO contact information if found.
    """
    try:
        device_id = request.headers.get("device-id", "")
        
        if not district:
            raise HTTPException(status_code=400, detail="District is required")
        
        # Query AEO table for active AEO in the district
        conn = database.get_db_connection()
        c = conn.cursor()
        
        aeo = c.execute('''
            SELECT id, name, phone, district, region, email, profile_picture
            FROM aeo
            WHERE district = ? AND is_active = 1
            LIMIT 1
        ''', (district,)).fetchone()
        
        conn.close()
        
        if aeo:
            return {
                "found": True,
                "id": aeo['id'],
                "name": aeo['name'],
                "phone": aeo['phone'],
                "district": aeo['district'],
                "region": aeo['region'] if aeo['region'] else "N/A",
                "email": aeo['email'] if aeo['email'] else None,
                "profile_picture": aeo['profile_picture'] if aeo['profile_picture'] else None
            }
        else:
            return {
                "found": False,
                "message": "No Extension Officer assigned to this district yet"
            }
            
    except Exception as e:
        print(f"[API] Error fetching AEO: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch AEO information: {str(e)}")


@app.post("/api/farmer/support-request")
async def submit_support_request(
    request: Request,
    category: str = Form(...),
    subject: str = Form(...),
    message: str = Form(...),
    name: str = Form(None),
    phone: str = Form(None)
):
    """
    Submit a support request from a farmer.
    Saves to farmer_support_requests table for superadmin review.
    """
    try:
        device_id = request.headers.get("device-id", "")
        
        if not device_id:
            raise HTTPException(status_code=400, detail="Device ID is required")
        
        if not category or not subject or not message:
            raise HTTPException(status_code=400, detail="Category, subject, and message are required")
        
        # Save to database
        conn = database.get_db_connection()
        c = conn.cursor()
        
        c.execute('''
            INSERT INTO farmer_support_requests 
            (farmer_id, category, subject, message, name, phone, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 'pending', CURRENT_TIMESTAMP)
        ''', (device_id, category, subject, message, name, phone))
        
        request_id = c.lastrowid
        conn.commit()
        conn.close()
        
        return {
            "success": True,
            "request_id": request_id,
            "message": "Support request submitted successfully"
        }
        
    except Exception as e:
        print(f"[API] Error submitting support request: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to submit support request: {str(e)}")


@app.get("/api/superadmin/support-requests")
async def get_support_requests(
    request: Request,
    status_filter: str = Query(None, description="Filter by status: pending, resolved, all")
):
    """
    Get all farmer support requests for superadmin dashboard.
    """
    try:
        # TODO: Add authentication check for superadmin
        
        conn = database.get_db_connection()
        c = conn.cursor()
        
        if status_filter and status_filter != "all":
            requests = c.execute('''
                SELECT fsr.*, f.name as farmer_name, f.phone as farmer_phone
                FROM farmer_support_requests fsr
                LEFT JOIN farmers f ON fsr.farmer_id = f.device_id
                WHERE fsr.status = ?
                ORDER BY fsr.created_at DESC
            ''', (status_filter,)).fetchall()
        else:
            requests = c.execute('''
                SELECT fsr.*, f.name as farmer_name, f.phone as farmer_phone
                FROM farmer_support_requests fsr
                LEFT JOIN farmers f ON fsr.farmer_id = f.device_id
                ORDER BY fsr.created_at DESC
            ''').fetchall()
        
        conn.close()
        
        return {
            "success": True,
            "requests": [dict(row) for row in requests]
        }
        
    except Exception as e:
        print(f"[API] Error fetching support requests: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch support requests: {str(e)}")


# ============================================================================
# AEO API ENDPOINTS - Profile Management
# ============================================================================

@app.post("/api/aeo/update-profile")
async def update_aeo_profile(
    request: Request,
    profile_picture: str = Form(None),
    phone: str = Form(None),
    district: str = Form(None),
    region: str = Form(None),
    email: str = Form(None)
):
    """
    Update AEO profile information.
    Requires AEO to be logged in (checks session).
    """
    try:
        # Get AEO ID from session
        aeo_id = request.session.get("aeo_id")
        if not aeo_id:
            raise HTTPException(status_code=401, detail="Not authenticated")
        
        # Validate at least one field is provided
        if not any([profile_picture, phone, district, region, email]):
            raise HTTPException(status_code=400, detail="At least one field must be provided")
        
        # Update database
        conn = database.get_db_connection()
        c = conn.cursor()
        
        update_fields = []
        update_values = []
        
        if profile_picture:
            update_fields.append("profile_picture = ?")
            update_values.append(profile_picture)
        
        if phone:
            update_fields.append("phone = ?")
            update_values.append(phone)
        
        if district:
            update_fields.append("district = ?")
            update_values.append(district)
        
        if region:
            update_fields.append("region = ?")
            update_values.append(region)
        
        if email:
            update_fields.append("email = ?")
            update_values.append(email)
        
        update_values.append(aeo_id)
        
        query = f"UPDATE aeo SET {', '.join(update_fields)} WHERE id = ?"
        c.execute(query, tuple(update_values))
        
        conn.commit()
        conn.close()
        
        return {
            "success": True,
            "message": "Profile updated successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[API] Error updating AEO profile: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update profile: {str(e)}")


if __name__ == "__main__":
    uvicorn.run("main:app", host=HOST, port=PORT, reload=DEBUG)
